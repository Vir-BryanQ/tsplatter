import os
import gc
import sys
import torch
import random
import json
import shutil
import subprocess
import argparse
import pandas as pd
import signal
import time
from datetime import datetime
import psutil
import uuid
from tqdm import tqdm
import numpy as np

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(parent_dir)

from utils.general_utils import safe_state
from tsplatter import preallocate_vmem, release_vmem
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

def generate_configs(args):
    clustering_threshold_values = np.arange(0.7, 1 + 1e-6, 0.01)
    use_max_weight_values = [True, False]
    max_threshold_values = np.arange(0.01, 0.5 + 1e-6, 0.01)
    sum_threshold_values = np.arange(0.1, 5.0 + 1e-6, 0.1)

    # 固定超参数
    fixed_params = {
        "k_neighbors": 100,
        "lambda_reg": 1e-3,
        "pow_k": 2,
        "topk": 45,
        "feature_level": 3,
        "encoder": "dino",
    }

    all_configs = []

    for clustering_threshold in clustering_threshold_values:
        for use_max_weight in use_max_weight_values:

            if use_max_weight:
                for max_threshold in max_threshold_values:
                    config = {
                        "clustering_threshold": clustering_threshold,
                        "use_max_weight": use_max_weight,
                        "max_threshold": max_threshold,
                        "sum_threshold": 999,  # 不用
                        **fixed_params
                    }
                    all_configs.append(config)
            else:
                for sum_threshold in sum_threshold_values:
                    config = {
                        "clustering_threshold": clustering_threshold,
                        "use_max_weight": use_max_weight,
                        "max_threshold": 999,  # 不用
                        "sum_threshold": sum_threshold,
                        **fixed_params
                    }
                    all_configs.append(config)

    return all_configs

def append_row(output_excel, row_values):
    wb = load_workbook(output_excel)
    ws = wb.active
    ws.append(row_values)
    wb.save(output_excel)

def parse_arguments():
    parser = argparse.ArgumentParser(description='Training and evaluation loop')
    parser.add_argument('--dataset_path', type=str, required=True, help='Path to the dataset')
    parser.add_argument('--checkpoint_dir', type=str, required=True, help='Number of loops to run')
    parser.add_argument('--output_excel', type=str, required=True, help='Name of the output Excel file')
    parser.add_argument('--metric_json', type=str, required=True, help='Path to the metric JSON file')
    parser.add_argument('--scene_name', type=str, required=True, help='Name of the scene')
    # parser.add_argument('--tmp', type=float, required=True)
    return parser.parse_args()

def perform_sampling(args):
    preallocate_vmem()

    if not os.path.exists(args.output_excel):
        wb = Workbook()
        ws = wb.active
        ws.append([
            'psnr in paper', 'ssim in paper', 'lpips in paper',
            'psnr_rep', 'ssim_rep', 'lpips_rep',
            'psnr', 'ssim', 'lpips', 'psnr1', 'ssim1', 'lpips1', 
            'delta_psnr', 'delta_ssim', 'delta_lpips', 
            'clustering_threshold',	
            'use_max_weight',
            'max_threshold',
            'sum_threshold',
            'k_neighbors"',
            'lambda_reg',
            'pow_k',
            'topk',
            'feature_level',
            'encoder'
        ])
        bold_font = Font(bold=True)
        for cell in ws[1]: 
            cell.font = bold_font
        wb.save(args.output_excel)

    renders_dir = os.path.join(args.checkpoint_dir, 'renders')
    os.makedirs(renders_dir, exist_ok=True)
    evals_dir = os.path.join(args.checkpoint_dir, 'evals')
    os.makedirs(evals_dir, exist_ok=True)

    with open(args.metric_json, 'r') as f:
        metric = json.load(f)
    psnr_in_paper = metric['result'][args.scene_name]['psnr']
    ssim_in_paper = metric['result'][args.scene_name]['ssim']
    lpips_in_paper = metric['result'][args.scene_name]['lpips']

    psnr_rep = metric['result1'][args.scene_name]['psnr']
    ssim_rep = metric['result1'][args.scene_name]['ssim']
    lpips_rep = metric['result1'][args.scene_name]['lpips']

    eval_json = os.path.join(args.checkpoint_dir, 'eval.json')
    with open(eval_json, 'r') as f:
        eval_results = json.load(f)
    psnr = eval_results['results']['psnr']
    ssim = eval_results['results']['ssim']
    lpips = eval_results['results']['lpips']

    all_configs = generate_configs(args)

    checkpoint_name = os.listdir(os.path.join(args.checkpoint_dir, 'nerfstudio_models'))[0]
    checkpoint_path = os.path.join(args.checkpoint_dir, 'nerfstudio_models', checkpoint_name) 
    wait_list = [7, 7, 3, 5]
    for loop in tqdm(range(len(all_configs))):
        cfg = all_configs[loop]
        if cfg['use_max_weight']:
            smoothing_command = (f"python smoothing.py -s {args.dataset_path} --start_checkpoint {checkpoint_path} "
                                f"--clustering_threshold {cfg['clustering_threshold']} "	
                                f"--use_max_weight "
                                f"--max_threshold {cfg['max_threshold']} "
                                f"--sum_threshold {cfg['sum_threshold']} "
                                f"--k_neighbors {cfg['k_neighbors']} "
                                f"--lambda_reg {cfg['lambda_reg']} "
                                f"--pow_k {cfg['pow_k']} "
                                f"--topk {cfg['topk']} "
                                f"--feature_level {cfg['feature_level']} "
                                f"--encoder {cfg['encoder']}")
        else:
            smoothing_command = (f"python smoothing.py -s {args.dataset_path} --start_checkpoint {checkpoint_path} "
                                f"--clustering_threshold {cfg['clustering_threshold']} "	
                                f"--max_threshold {cfg['max_threshold']} "
                                f"--sum_threshold {cfg['sum_threshold']} "
                                f"--k_neighbors {cfg['k_neighbors']} "
                                f"--lambda_reg {cfg['lambda_reg']} "
                                f"--pow_k {cfg['pow_k']} "
                                f"--topk {cfg['topk']} "
                                f"--feature_level {cfg['feature_level']} "
                                f"--encoder {cfg['encoder']}")
        print(smoothing_command)
        process = subprocess.Popen(smoothing_command, shell=True, env=os.environ)

        time.sleep(wait_list[2])
        release_vmem()
        process.wait()

        preallocate_vmem()

        # Execute evaluation command again for smoothed results
        eval_command = (f"ns-eval --load-config {os.path.join(args.checkpoint_dir, 'config.yml')} "
            f"--output-path {os.path.join(evals_dir, f'eval{loop}.json')} "
            f"--render-output-path {os.path.join(renders_dir, f'render{loop}')}")
        print(eval_command)
        process = subprocess.Popen(eval_command, shell=True, env=os.environ)

        time.sleep(wait_list[3])
        release_vmem()
        process.wait()

        preallocate_vmem()

        eval1_json_path = os.path.join(evals_dir, f'eval{loop}.json')
        with open(eval1_json_path, 'r') as f:
            eval1_results = json.load(f)

        psnr1 = eval1_results['results']['psnr']
        ssim1 = eval1_results['results']['ssim']
        lpips1 = eval1_results['results']['lpips']

        # Calculate differences
        delta_psnr = psnr1 - psnr
        delta_ssim = ssim1 - ssim
        delta_lpips = lpips1 - lpips

        print(f"psnr_in_paper={psnr_in_paper}")
        print(f"ssim_in_paper={ssim_in_paper}")
        print(f"lpips_in_paper={lpips_in_paper}")
        
        print(f"psnr_rep={psnr_rep}")
        print(f"ssim_rep={ssim_rep}")
        print(f"lpips_rep={lpips_rep}")

        print(f"psnr={psnr}")
        print(f"ssim={ssim}")
        print(f"lpips={lpips}")

        print(f"psnr1={psnr1}")
        print(f"ssim1={ssim1}")
        print(f"lpips1={lpips1}")

        print(f"delta_psnr={delta_psnr}")
        print(f"delta_ssim={delta_ssim}")
        print(f"delta_lpips={delta_lpips}")

        print(f"clustering_threshold={cfg['clustering_threshold']}")	
        print(f"use_max_weight={cfg['use_max_weight']}")
        print(f"max_threshold={cfg['max_threshold']}")
        print(f"sum_threshold={cfg['sum_threshold']}")
        print(f"k_neighbors={cfg['k_neighbors']}")
        print(f"lambda_reg={cfg['lambda_reg']}")
        print(f"pow_k={cfg['pow_k']}")
        print(f"topk={cfg['topk']}")
        print(f"feature_level={cfg['feature_level']}")
        print(f"encoder={cfg['encoder']}")

        append_row(args.output_excel, [
            psnr_in_paper, ssim_in_paper, lpips_in_paper,
            psnr_rep, ssim_rep, lpips_rep,
            psnr, ssim, lpips,
            psnr1, ssim1, lpips1,
            delta_psnr, delta_ssim, delta_lpips,
            cfg['clustering_threshold'],
            cfg['use_max_weight'],
            cfg['max_threshold'],
            cfg['sum_threshold'],
            cfg['k_neighbors'],
            cfg['lambda_reg'],
            cfg['pow_k'],
            cfg['topk'],
            cfg['feature_level'],
            cfg['encoder']
            ])

        print(f'Row-{loop} in {args.output_excel} saved.')


if __name__ == '__main__':
    args = parse_arguments()
    safe_state(False)
    perform_sampling(args)
